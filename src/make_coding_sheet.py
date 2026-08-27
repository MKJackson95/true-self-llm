"""Generate a stratified sample for blind human coding.

    python src/make_coding_sheet.py

Writes data/coding/blind_coding_sheet.xlsx — a workbook containing 400
explanations sampled evenly across model, item class and version, with empty
columns for the twelve codes.

The sheet deliberately omits the automated codes, the rating, the forced
choice, and the model that produced the response. A human coder who can see
any of those is not coding independently, and the agreement statistic would
measure willingness to confirm rather than agreement about the text.

The mapping from row number back to call_id is written separately to
data/coding/blind_key.csv, which should not be opened until coding is
complete.
"""

from __future__ import annotations

import argparse
import csv
import json
import random
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CODES = [
    ("MORAL_VALENCE", "The behaviour is good (or bad), therefore it does (or does not) reflect the true self."),
    ("REFLECTIVE", "Appeals to deliberation, considered judgement, or the agent's endorsed beliefs against their impulses."),
    ("EMOTIONAL", "Appeals to urges, cravings or feelings as more revealing than considered belief."),
    ("META_DESIRE", "Appeals to whether the agent wants to have the motivation they have."),
    ("ESSENTIALISM", "Appeals to a fixed underlying nature or essence, prior to and independent of conduct."),
    ("PERSON_POSITIVITY", "Appeals to a general disposition to think well of people."),
    ("ATTITUDE_NOT_SELF", "Separates the agent's surface attitudes from a deeper self, allowing both at once."),
    ("ANTI_CONDITIONING", "Appeals to resisting social pressure or holding a view independently of one's surroundings."),
    ("CONSISTENCY", "Appeals to the agent's prior pattern or track record as evidence about their real character."),
    ("TRIVIALITY", "Declines to attribute because the domain is too superficial to bear on the true self."),
    ("AUTHENTICITY_BARE", "Asserts the behaviour is or is not genuinely the agent's own, giving no criterion."),
    ("GROWTH", "Treats change as making the self rather than revealing it."),
    ("UNCLEAR", "No criterion identifiable, or too brief to code."),
]

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
INPUT_FILL = PatternFill("solid", fgColor="FFFFCC")
FONT = "Arial"


def stratified_sample(rows: list[dict], n: int, seed: int) -> list[dict]:
    """Even coverage across model x item_class x version x measure."""
    rng = random.Random(seed)
    strata = defaultdict(list)
    for r in rows:
        strata[(r["model_key"], r["item_class"], r["version"], r["measure"])].append(r)

    keys = sorted(strata)
    per = max(1, n // len(keys))
    picked = []
    for k in keys:
        pool = strata[k]
        rng.shuffle(pool)
        picked.extend(pool[:per])

    remaining = [r for r in rows if r not in picked]
    rng.shuffle(remaining)
    picked.extend(remaining[: max(0, n - len(picked))])
    rng.shuffle(picked)
    return picked[:n]


def write_workbook(sample: list[dict], path: Path) -> None:
    wb = Workbook()

    # --- Instructions -----------------------------------------------------
    ws = wb.active
    ws.title = "Instructions"
    lines = [
        ("How to use this sheet", True),
        ("", False),
        ("Go to the Coding tab. Read each explanation and put an x in every column whose", False),
        ("criterion the explanation invokes. More than one may apply; that is expected and", False),
        ("normal. If none applies, mark UNCLEAR.", False),
        ("", False),
        ("Code what the explanation says, not what it implies. An explanation accompanying", False),
        ("a high rating for a good change is not MORAL_VALENCE unless the explanation itself", False),
        ("appeals to the goodness of the behaviour.", False),
        ("", False),
        ("Code the reasoning offered, not whether it is correct or consistent.", False),
        ("", False),
        ("Use the Notes column when a response does not fit cleanly, or straddles two codes.", False),
        ("Those cases are the most useful thing this exercise produces.", False),
        ("", False),
        ("The model that produced each response, its rating, and the automated coding are", False),
        ("deliberately withheld. Do not look them up before finishing.", False),
        ("", False),
        ("Code definitions", True),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name=FONT, bold=bold, size=12 if bold else 11)

    start = len(lines) + 1
    for j, (code, desc) in enumerate(CODES):
        ws.cell(row=start + j, column=1, value=code).font = Font(name=FONT, bold=True)
        c = ws.cell(row=start + j, column=2, value=desc)
        c.font = Font(name=FONT)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95

    # --- Coding -----------------------------------------------------------
    cs = wb.create_sheet("Coding")
    headers = ["row", "explanation"] + [c for c, _ in CODES] + ["notes"]
    for j, h in enumerate(headers, start=1):
        c = cs.cell(row=1, column=j, value=h)
        c.font = Font(name=FONT, bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="bottom",
                                textRotation=90 if j > 2 and h != "notes" else 0)

    for i, r in enumerate(sample, start=2):
        cs.cell(row=i, column=1, value=i - 1).font = Font(name=FONT)
        e = cs.cell(row=i, column=2, value=r["explanation"])
        e.font = Font(name=FONT)
        e.alignment = Alignment(wrap_text=True, vertical="top")
        for j in range(len(CODES)):
            cell = cs.cell(row=i, column=3 + j)
            cell.fill = INPUT_FILL
            cell.alignment = Alignment(horizontal="center")
        n = cs.cell(row=i, column=3 + len(CODES))
        n.fill = INPUT_FILL
        n.alignment = Alignment(wrap_text=True, vertical="top")
        cs.row_dimensions[i].height = 60

    cs.column_dimensions["A"].width = 6
    cs.column_dimensions["B"].width = 90
    for j in range(len(CODES)):
        cs.column_dimensions[get_column_letter(3 + j)].width = 5
    cs.column_dimensions[get_column_letter(3 + len(CODES))].width = 40
    cs.freeze_panes = "C2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--processed", default="data/processed/nbk_study1.csv")
    ap.add_argument("--out", default="data/coding/blind_coding_sheet.xlsx")
    ap.add_argument("--key", default="data/coding/blind_key.csv")
    ap.add_argument("--n", type=int, default=400)
    ap.add_argument("--seed", type=int, default=20260826)
    args = ap.parse_args(argv)

    rows = [r for r in csv.DictReader(open(args.processed, encoding="utf-8"))
            if r["parse_status"] in ("ok", "ok_word") and r["explanation"].strip()]

    sample = stratified_sample(rows, args.n, args.seed)
    write_workbook(sample, Path(args.out))

    key_path = Path(args.key)
    with key_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["row", "call_id", "model_key", "item_id", "item_class",
                    "version", "measure", "frame"])
        for i, r in enumerate(sample, start=1):
            w.writerow([i, r["call_id"], r["model_key"], r["item_id"],
                        r["item_class"], r["version"], r["measure"], r["frame"]])

    strata = defaultdict(int)
    for r in sample:
        strata[(r["model_key"], r["item_class"])] += 1
    print(f"{len(sample)} responses written to {args.out}")
    print(f"key written to {key_path} — do not open until coding is complete\n")
    print("sample composition:")
    for k in sorted(strata):
        print(f"  {k[0]:9}{k[1]:12}{strata[k]:4}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
