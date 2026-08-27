"""Compare blind human coding against the automated coding.

    python src/coding_agreement.py

Joins the completed blind coding sheet to the automated codes via the key,
and reports agreement per code and overall.

Reliability is reported as Krippendorff's alpha computed separately for each
code, treating each code as a binary decision across the sampled responses.
This is the appropriate form for multi-label data: an overall alpha across a
label set would conflate codes that agree well with codes that do not, and it
is the per-code figure that tells you which definitions are working.

Alpha corrects for chance agreement, which matters here because several codes
are rare. Raw agreement of 0.95 on a code applied to 2% of responses is what
two coders achieve by both saying no, and alpha will show that as near zero.

Also reported, per code:

  human_only    the human marked it, the coder did not
  auto_only     the coder marked it, the human did not

The direction of disagreement is diagnostic. A code the automated coder
applies four times as often as the human does is not noisy; it is being read
differently, and the definition is the thing to fix.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

CODES = [
    "MORAL_VALENCE", "REFLECTIVE", "EMOTIONAL", "META_DESIRE", "ESSENTIALISM",
    "PERSON_POSITIVITY", "ATTITUDE_NOT_SELF", "ANTI_CONDITIONING",
    "CONSISTENCY", "TRIVIALITY", "AUTHENTICITY_BARE", "GROWTH", "UNCLEAR",
]


def krippendorff_binary(pairs: list[tuple[int, int]]) -> float:
    """Krippendorff's alpha for two coders on a binary variable.

    With two coders and no missing data this reduces to a form computable
    directly from the observed and expected disagreement. Returns nan where
    every value is identical across both coders and all units, since alpha is
    undefined when there is no variation to explain.
    """
    n = len(pairs)
    if n == 0:
        return float("nan")

    # Observed disagreement: proportion of units where the two coders differ.
    do = sum(1 for a, b in pairs if a != b) / n

    # Expected disagreement under independence, from the pooled marginal.
    ones = sum(a + b for a, b in pairs)
    total = 2 * n
    p1 = ones / total
    p0 = 1 - p1
    if p1 in (0.0, 1.0):
        return float("nan")

    # Expected disagreement from the coincidence matrix, normalised over all
    # pairable values: de = 2 * n0 * n1 / (n * (n - 1)).
    n0, n1 = total * p0, total * p1
    de = (2 * n0 * n1) / (total * (total - 1))
    if de == 0:
        return float("nan")
    return 1 - (do / de)


def load_human(path: Path) -> dict[int, set[str]]:
    wb = load_workbook(path)
    ws = wb["Coding"]
    header = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
    idx = {c: header.index(c) + 1 for c in CODES}

    out: dict[int, set[str]] = {}
    for i in range(2, ws.max_row + 1):
        row_no = ws.cell(row=i, column=1).value
        if row_no is None:
            continue
        marked = {c for c in CODES
                  if str(ws.cell(row=i, column=idx[c]).value or "").strip() != ""}
        out[int(row_no)] = marked
    return out


def load_key(path: Path) -> dict[int, dict]:
    with path.open(encoding="utf-8") as fh:
        return {int(r["row"]): r for r in csv.DictReader(fh)}


def load_auto(path: Path) -> dict[str, set[str]]:
    out: dict[str, set[str]] = {}
    for line in path.open(encoding="utf-8"):
        line = line.strip()
        if line:
            rec = json.loads(line)
            out[rec["call_id"]] = set(rec["codes"])
    return out


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--human", default="data/coding/blind_coding_sheet_completed.xlsx")
    ap.add_argument("--key", default="data/coding/blind_key.csv")
    ap.add_argument("--auto", default="data/raw/codes_auto.jsonl")
    ap.add_argument("--out", default="data/processed/coding_agreement.csv")
    args = ap.parse_args(argv)

    human = load_human(Path(args.human))
    key = load_key(Path(args.key))
    auto = load_auto(Path(args.auto))

    joined = []
    missing = 0
    for row_no, hcodes in human.items():
        k = key.get(row_no)
        if not k or k["call_id"] not in auto:
            missing += 1
            continue
        joined.append((row_no, k, hcodes, auto[k["call_id"]]))

    print(f"{len(joined)} responses joined"
          + (f", {missing} could not be matched" if missing else ""))

    # --- per-code agreement ------------------------------------------------
    results = []
    for code in CODES:
        pairs = [(int(code in h), int(code in a)) for _, _, h, a in joined]
        both = sum(1 for x, y in pairs if x and y)
        h_only = sum(1 for x, y in pairs if x and not y)
        a_only = sum(1 for x, y in pairs if y and not x)
        neither = sum(1 for x, y in pairs if not x and not y)
        raw = (both + neither) / len(pairs)
        alpha = krippendorff_binary(pairs)
        results.append({
            "code": code,
            "human_n": both + h_only,
            "auto_n": both + a_only,
            "both": both,
            "human_only": h_only,
            "auto_only": a_only,
            "raw_agreement": round(raw, 3),
            "alpha": round(alpha, 3) if alpha == alpha else "",
        })

    print(f"\n{'code':20}{'human':>7}{'auto':>7}{'both':>7}{'h only':>8}"
          f"{'a only':>8}{'raw':>7}{'alpha':>8}")
    print("-" * 72)
    for r in sorted(results, key=lambda x: -x["human_n"]):
        a = r["alpha"] if r["alpha"] != "" else "  n/a"
        print(f"{r['code']:20}{r['human_n']:7}{r['auto_n']:7}{r['both']:7}"
              f"{r['human_only']:8}{r['auto_only']:8}{r['raw_agreement']:7.2f}"
              f"{a:>8}")

    # --- overall -----------------------------------------------------------
    all_pairs = [(int(c in h), int(c in a))
                 for _, _, h, a in joined for c in CODES]
    overall_raw = sum(1 for x, y in all_pairs if x == y) / len(all_pairs)
    overall_alpha = krippendorff_binary(all_pairs)

    exact = sum(1 for _, _, h, a in joined if h == a)
    overlap = sum(1 for _, _, h, a in joined if h & a)

    print(f"\nAcross all code decisions: raw agreement {overall_raw:.3f}, "
          f"alpha {overall_alpha:.3f}")
    print(f"Identical code sets: {exact}/{len(joined)} ({exact/len(joined):.1%})")
    print(f"At least one code in common: {overlap}/{len(joined)} "
          f"({overlap/len(joined):.1%})")

    # --- where the human found more or fewer codes -------------------------
    hn = sum(len(h) for _, _, h, _ in joined) / len(joined)
    an = sum(len(a) for _, _, _, a in joined) / len(joined)
    print(f"\nMean codes per response: human {hn:.2f}, automated {an:.2f}")

    # --- by item class, since the sample is not compositionally matched ----
    print("\nMean codes per response by item class:")
    by = defaultdict(lambda: [0, 0, 0])
    for _, k, h, a in joined:
        b = by[k["item_class"]]
        b[0] += len(h); b[1] += len(a); b[2] += 1
    for cls, (hh, aa, n) in sorted(by.items()):
        print(f"  {cls:12} human {hh/n:.2f}   automated {aa/n:.2f}   (n={n})")

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(results[0]))
        w.writeheader()
        w.writerows(results)
    print(f"\nper-code results written to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
